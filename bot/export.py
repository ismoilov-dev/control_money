"""Excel export generator for FinMate Bot."""

from __future__ import annotations

import io
from typing import Any
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def generate_excel_export(transactions: list[dict[str, Any]]) -> bytes:
    """Generate Excel file (.xlsx) bytes from transactions list."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Xarajatlar"

    # Header styling
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    headers = ["ID", "Sana (UTC)", "Turi", "Kategoriya", "Summa (so'm)", "Izoh"]
    ws.append(headers)

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center

    for r in transactions:
        tx_type = "Kirim" if r.get("type") == "income" else "Xarajat"
        ws.append([
            r.get("id"),
            r.get("created_at"),
            tx_type,
            r.get("category", ""),
            r.get("amount", 0),
            r.get("description", ""),
        ])

    # Format data rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
        row[0].alignment = align_center
        row[1].alignment = align_center
        row[2].alignment = align_center
        row[3].alignment = align_left
        row[4].alignment = align_right
        row[4].number_format = "#,##0"
        row[5].alignment = align_left

    # Auto adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
